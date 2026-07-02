# backend/routes/super_admin.py
# Dashboard pemilik platform — kelola sekolah (tenant), status pembayaran,
# dan domain email per sekolah. Endpoint ini SENGAJA tidak diturunkan dari
# require_admin (lihat auth.py require_super_admin) supaya admin sekolah
# biasa tidak pernah bisa naik privilege ke sini walau ada bug di tempat
# lain.
from flask import Blueprint, request, jsonify, send_file
import re, uuid, io
from datetime import date, timedelta
from db import query, delete_school_cascade
from auth import require_super_admin
import storage

_FEATURE_FIELDS = ['feature_export', 'feature_bank_soal', 'feature_upload_media', 'feature_mobile']

super_admin_bp = Blueprint('super_admin', __name__)

_SLUG_RE = re.compile(r'[^a-z0-9]+')


def _slugify(name):
    base = _SLUG_RE.sub('-', name.strip().lower()).strip('-') or 'sekolah'
    slug = base
    n = 1
    while query("SELECT 1 FROM schools WHERE slug=%s", (slug,), fetch='one'):
        n += 1
        slug = f"{base}-{n}"
    return slug


def _default_paid_until(plan):
    """Kalau admin tidak menentukan tanggal lisensi: trial otomatis 14 hari,
    plan lain (standard/pro) otomatis 30 hari dari sekarang — supaya tiap
    sekolah selalu punya tanggal berlaku untuk ditampilkan countdown-nya."""
    days = 14 if plan == 'trial' else 30
    return (date.today() + timedelta(days=days)).isoformat()


@super_admin_bp.route('/api/super-admin/overview', methods=['GET'])
@require_super_admin
def overview():
    """Ringkasan lintas-sekolah untuk kartu statistik di atas tabel — guru/
    siswa per sekolah sudah ada di baris tabelnya masing-masing, jadi di
    sini cuma angka yang genuinely lintas-sekolah: total sekolah, total
    ujian, beban real-time (sesi ujian yang sedang berjalan SEKARANG), dan
    pemakaian Supabase Storage."""
    row = query("""
        SELECT
            (SELECT COUNT(*) FROM schools) as total_schools,
            (SELECT COUNT(*) FROM schools WHERE is_active=true) as active_schools,
            (SELECT COUNT(*) FROM exams) as total_exams,
            (SELECT COUNT(*) FROM exam_sessions WHERE submitted_at IS NULL AND status='ongoing') as active_sessions
    """, fetch='one')
    result = dict(row)
    result['storage_bytes'] = storage.get_storage_usage()
    return jsonify(result)


@super_admin_bp.route('/api/super-admin/schools', methods=['GET'])
@require_super_admin
def list_schools():
    rows = query("""
        SELECT s.*,
               (SELECT COUNT(*) FROM users u WHERE u.school_id=s.id AND u.role='guru') as guru_count,
               (SELECT COUNT(*) FROM users u WHERE u.school_id=s.id AND u.role='siswa') as siswa_count,
               (SELECT COUNT(*) FROM exams e WHERE e.school_id=s.id) as exam_count,
               (SELECT COUNT(*) FROM exam_sessions es WHERE es.school_id=s.id
                  AND es.submitted_at IS NULL AND es.status='ongoing') as active_sessions
        FROM schools s
        ORDER BY s.created_at DESC
    """)
    return jsonify([dict(r) for r in rows])


@super_admin_bp.route('/api/super-admin/schools', methods=['POST'])
@require_super_admin
def create_school():
    data = request.json or {}
    name = (data.get('name') or '').strip()
    if not name:
        return jsonify({'error': 'Nama sekolah wajib'}), 400
    allowed_domain = (data.get('allowed_domain') or '').strip().lower() or None
    plan = (data.get('plan') or 'standard').strip()
    paid_until = data.get('paid_until') or _default_paid_until(plan)
    slug = _slugify(name)
    school_id = str(uuid.uuid4())
    query("""INSERT INTO schools (id, name, slug, plan, is_active, paid_until, allowed_domain)
             VALUES (%s,%s,%s,%s,true,%s,%s)""",
          (school_id, name, slug, plan, paid_until, allowed_domain), fetch='none')
    row = query("SELECT * FROM schools WHERE id=%s", (school_id,), fetch='one')
    return jsonify(dict(row)), 201


@super_admin_bp.route('/api/super-admin/schools/<school_id>', methods=['PATCH'])
@require_super_admin
def update_school(school_id):
    data = request.json or {}
    fields, vals = [], []
    if 'name' in data:
        fields.append('name=%s'); vals.append((data['name'] or '').strip())
    if 'allowed_domain' in data:
        fields.append('allowed_domain=%s')
        vals.append((data['allowed_domain'] or '').strip().lower() or None)
    if 'plan' in data:
        fields.append('plan=%s'); vals.append((data['plan'] or 'standard').strip())
    if 'is_active' in data:
        fields.append('is_active=%s'); vals.append(bool(data['is_active']))
    if 'paid_until' in data:
        paid_until = data['paid_until'] or None
        if not paid_until:
            # Tanggal dikosongkan — default berdasarkan plan yang berlaku
            # (plan baru dari request ini kalau ikut diubah, atau plan
            # sekolah saat ini kalau tidak).
            plan_for_default = (data.get('plan') or '').strip()
            if not plan_for_default:
                existing = query("SELECT plan FROM schools WHERE id=%s", (school_id,), fetch='one')
                plan_for_default = (existing or {}).get('plan') or 'standard'
            paid_until = _default_paid_until(plan_for_default)
        fields.append('paid_until=%s'); vals.append(paid_until)
    for f in _FEATURE_FIELDS:
        if f in data:
            fields.append(f'{f}=%s'); vals.append(bool(data[f]))
    if not fields:
        return jsonify({'error': 'Tidak ada field yang valid'}), 400
    query(f"UPDATE schools SET {', '.join(fields)} WHERE id=%s", vals + [school_id], fetch='none')
    row = query("SELECT * FROM schools WHERE id=%s", (school_id,), fetch='one')
    if not row:
        return jsonify({'error': 'Sekolah tidak ditemukan'}), 404
    return jsonify(dict(row))


@super_admin_bp.route('/api/super-admin/schools/<school_id>', methods=['GET'])
@require_super_admin
def get_school(school_id):
    row = query("SELECT * FROM schools WHERE id=%s", (school_id,), fetch='one')
    if not row:
        return jsonify({'error': 'Sekolah tidak ditemukan'}), 404
    return jsonify(dict(row))


@super_admin_bp.route('/api/super-admin/schools/<school_id>', methods=['DELETE'])
@require_super_admin
def delete_school(school_id):
    """Hapus sekolah PERMANEN beserta seluruh datanya (lihat
    db.delete_school_cascade). Wajib kirim {"confirm_name": "<nama persis>"}
    di body — dicek di server (bukan cuma di UI) supaya endpoint ini tidak
    bisa kepicu tanpa sengaja walau lewat panggilan API langsung."""
    school = query("SELECT id, name FROM schools WHERE id=%s", (school_id,), fetch='one')
    if not school:
        return jsonify({'error': 'Sekolah tidak ditemukan'}), 404
    confirm = ((request.json or {}).get('confirm_name') or '').strip()
    if confirm != school['name']:
        return jsonify({'error': 'Nama konfirmasi tidak cocok. Hapus dibatalkan.'}), 400
    delete_school_cascade(school_id)
    return jsonify({'ok': True})


@super_admin_bp.route('/api/super-admin/schools/<school_id>/detail', methods=['GET'])
@require_super_admin
def school_detail(school_id):
    """Rincian lebih dalam untuk satu sekolah — dipakai panel dropdown di
    dashboard super admin: breakdown status ujian, guru pending, dan
    aktivitas terakhir."""
    school = query("SELECT * FROM schools WHERE id=%s", (school_id,), fetch='one')
    if not school:
        return jsonify({'error': 'Sekolah tidak ditemukan'}), 404
    exam_status = query("""
        SELECT status, COUNT(*) as n FROM exams WHERE school_id=%s GROUP BY status
    """, (school_id,))
    guru_pending = query("""
        SELECT COUNT(*) as n FROM users WHERE school_id=%s AND role='guru_pending'
    """, (school_id,), fetch='one')
    rooms_count = query("SELECT COUNT(*) as n FROM rooms WHERE school_id=%s", (school_id,), fetch='one')
    last_activity = query("""
        SELECT al.action, al.detail, al.created_at, u.name as user_name
        FROM activity_logs al LEFT JOIN users u ON u.id=al.user_id
        WHERE al.school_id=%s ORDER BY al.created_at DESC LIMIT 10
    """, (school_id,))
    return jsonify({
        'school': dict(school),
        'exam_status': {r['status']: r['n'] for r in exam_status},
        'guru_pending': guru_pending['n'],
        'rooms_count': rooms_count['n'],
        'recent_activity': [dict(r) for r in last_activity],
    })


@super_admin_bp.route('/api/super-admin/schools/<school_id>/monitor', methods=['GET'])
@require_super_admin
def school_monitor(school_id):
    """Pantau real-time sesi ujian yang sedang berlangsung di SATU sekolah —
    versi super_admin dari Monitor Ujian Aktif yang admin sekolah biasa
    sudah punya (lihat routes/admin.py exam_monitor), supaya super_admin
    bisa cek langsung tanpa perlu login sebagai admin sekolah itu."""
    school = query("SELECT id, name FROM schools WHERE id=%s", (school_id,), fetch='one')
    if not school:
        return jsonify({'error': 'Sekolah tidak ditemukan'}), 404
    rows = query("""
        SELECT es.id as session_id, es.exam_id, es.student_id, es.status,
               es.started_at, es.expires_at, es.tab_violations,
               es.ip_address,
               u.name as student_name, c.name as class_name,
               e.title as exam_title,
               (SELECT COUNT(*) FROM answers a WHERE a.session_id=es.id) as answered_count,
               (SELECT COUNT(*) FROM questions q WHERE q.exam_id=es.exam_id) as total_questions
        FROM exam_sessions es
        JOIN users u ON u.id = es.student_id
        LEFT JOIN classes c ON c.id = u.class_id
        JOIN exams e ON e.id = es.exam_id
        WHERE es.submitted_at IS NULL AND es.status='ongoing' AND es.school_id=%s
        ORDER BY es.started_at DESC
    """, (school_id,))
    return jsonify([dict(r) for r in rows])


# ── Admin sekolah ────────────────────────────────────────────────
# Admin sekolah biasa (require_admin) cuma bisa membuat user UNTUK
# sekolahnya sendiri — begitu sekolah baru dibuat lewat dashboard ini,
# belum ada satupun admin di sana, jadi tidak ada yang bisa login untuk
# membuat admin pertama (ayam-telur). Endpoint ini cuma untuk super_admin,
# khusus membuat/lihat akun role='admin' di sekolah manapun.
@super_admin_bp.route('/api/super-admin/schools/<school_id>/admins', methods=['GET'])
@require_super_admin
def list_school_admins(school_id):
    rows = query("""SELECT id, name, email, is_active, created_at
                     FROM users WHERE school_id=%s AND role='admin' ORDER BY created_at""",
                 (school_id,))
    return jsonify([dict(r) for r in rows])


@super_admin_bp.route('/api/super-admin/schools/<school_id>/admins', methods=['POST'])
@require_super_admin
def create_school_admin(school_id):
    from auth import hash_password, validate_email_domain
    school = query("SELECT id FROM schools WHERE id=%s", (school_id,), fetch='one')
    if not school:
        return jsonify({'error': 'Sekolah tidak ditemukan'}), 404

    data  = request.json or {}
    name  = (data.get('name') or '').strip()
    email = (data.get('email') or '').strip().lower()
    pw    = (data.get('password') or '').strip()
    if not name or not email:
        return jsonify({'error': 'Nama dan email wajib'}), 400
    if not pw or len(pw) < 6:
        return jsonify({'error': 'Password minimal 6 karakter'}), 400
    domain_err = validate_email_domain(school_id, email)
    if domain_err:
        return jsonify({'error': domain_err}), 400
    existing = query("SELECT id FROM users WHERE LOWER(email)=%s", (email,), fetch='one')
    if existing:
        return jsonify({'error': 'Email sudah terdaftar'}), 409

    uid = str(uuid.uuid4())
    dummy_gid = f"manual_{uuid.uuid4().hex[:12]}"
    query("""INSERT INTO users (id, google_id, email, name, role, password_hash, is_active, school_id)
             VALUES (%s,%s,%s,%s,'admin',%s,true,%s)""",
          (uid, dummy_gid, email, name, hash_password(pw), school_id), fetch='none')
    user = query("SELECT id, name, email, role, is_active, created_at FROM users WHERE id=%s", (uid,), fetch='one')
    return jsonify(dict(user)), 201


@super_admin_bp.route('/api/super-admin/schools/<school_id>/admins/<admin_id>', methods=['PATCH'])
@require_super_admin
def update_school_admin(school_id, admin_id):
    """Admin sekolah tidak bisa diedit oleh sesama admin (lihat admin.py) —
    ini jalur resminya, khusus Super Admin, untuk ubah nama/email/status
    aktif satu akun admin di sekolah tertentu."""
    admin = query("SELECT id FROM users WHERE id=%s AND school_id=%s AND role='admin'",
                 (admin_id, school_id), fetch='one')
    if not admin:
        return jsonify({'error': 'Admin tidak ditemukan'}), 404
    data = request.json or {}
    fields, vals = [], []
    if 'name' in data:
        name = (data['name'] or '').strip()
        if not name:
            return jsonify({'error': 'Nama tidak boleh kosong'}), 400
        fields.append('name=%s'); vals.append(name)
    if 'email' in data:
        email = (data['email'] or '').strip().lower()
        if not email:
            return jsonify({'error': 'Email tidak boleh kosong'}), 400
        existing = query("SELECT id FROM users WHERE LOWER(email)=%s AND id!=%s", (email, admin_id), fetch='one')
        if existing:
            return jsonify({'error': 'Email sudah dipakai akun lain'}), 409
        fields.append('email=%s'); vals.append(email)
    if 'is_active' in data:
        fields.append('is_active=%s'); vals.append(bool(data['is_active']))
    if not fields:
        return jsonify({'error': 'Tidak ada field yang diubah'}), 400
    query(f"UPDATE users SET {', '.join(fields)} WHERE id=%s", vals + [admin_id], fetch='none')
    user = query("SELECT id, name, email, role, is_active, created_at FROM users WHERE id=%s", (admin_id,), fetch='one')
    return jsonify(dict(user))


@super_admin_bp.route('/api/super-admin/schools/<school_id>/admins/<admin_id>/reset-password', methods=['POST'])
@require_super_admin
def reset_school_admin_password(school_id, admin_id):
    from auth import hash_password
    admin = query("SELECT id FROM users WHERE id=%s AND school_id=%s AND role='admin'",
                 (admin_id, school_id), fetch='one')
    if not admin:
        return jsonify({'error': 'Admin tidak ditemukan'}), 404
    pw = (request.json or {}).get('password', '').strip()
    if not pw or len(pw) < 6:
        return jsonify({'error': 'Password minimal 6 karakter'}), 400
    query("UPDATE users SET password_hash=%s WHERE id=%s", (hash_password(pw), admin_id), fetch='none')
    return jsonify({'ok': True})


@super_admin_bp.route('/api/super-admin/schools/<school_id>/admins/<admin_id>', methods=['DELETE'])
@require_super_admin
def delete_school_admin(school_id, admin_id):
    """Ini satu-satunya jalur untuk hapus akun admin (admin sekolah sengaja
    tidak boleh saling hapus, lihat admin.py delete_user). Ditolak kalau
    ini admin TERAKHIR di sekolah itu — jangan sampai sekolah kehilangan
    akses admin sama sekali."""
    admin = query("SELECT id, name FROM users WHERE id=%s AND school_id=%s AND role='admin'",
                 (admin_id, school_id), fetch='one')
    if not admin:
        return jsonify({'error': 'Admin tidak ditemukan'}), 404
    remaining = query("SELECT COUNT(*) as n FROM users WHERE school_id=%s AND role='admin' AND id!=%s",
                      (school_id, admin_id), fetch='one')['n']
    if remaining == 0:
        return jsonify({'error': 'Tidak bisa hapus — ini satu-satunya admin sekolah ini. Buat admin lain dulu.'}), 400
    query("DELETE FROM users WHERE id=%s", (admin_id,), fetch='none')
    return jsonify({'ok': True})


# ── Audit Log lintas sekolah ────────────────────────────────────
def _activity_log_filters():
    """Bangun WHERE + params dari query string, dipakai bareng oleh endpoint
    list (dipaginasi) dan export (tanpa limit) supaya filter selalu konsisten."""
    where, params = ["1=1"], []
    school_id = request.args.get('school_id', '').strip()
    action    = request.args.get('action', '').strip()
    role      = request.args.get('role', '').strip()
    q         = request.args.get('q', '').strip()
    date_from = request.args.get('date_from', '').strip()
    date_to   = request.args.get('date_to', '').strip()
    if school_id:
        where.append("al.school_id=%s"); params.append(school_id)
    if action:
        where.append("al.action=%s"); params.append(action)
    if role:
        where.append("u.role=%s"); params.append(role)
    if q:
        where.append("(al.detail ILIKE %s OR u.name ILIKE %s OR u.email ILIKE %s)")
        params += [f"%{q}%", f"%{q}%", f"%{q}%"]
    if date_from:
        where.append("al.created_at >= %s"); params.append(date_from)
    if date_to:
        where.append("al.created_at < (%s::date + INTERVAL '1 day')"); params.append(date_to)
    return " AND ".join(where), params


@super_admin_bp.route('/api/super-admin/activity-logs', methods=['GET'])
@require_super_admin
def list_activity_logs():
    """Audit log lintas sekolah — bisa difilter per sekolah, aksi, peran user,
    kata kunci, dan rentang tanggal. Dipaginasi karena bisa tumbuh besar
    (tiap login/buat ujian/publish/dsb tercatat lewat log_activity())."""
    where, params = _activity_log_filters()
    page      = max(1, int(request.args.get('page', 1) or 1))
    page_size = min(200, max(1, int(request.args.get('page_size', 50) or 50)))
    offset    = (page - 1) * page_size

    total = query(f"""
        SELECT COUNT(*) as n FROM activity_logs al
        LEFT JOIN users u ON u.id=al.user_id
        WHERE {where}
    """, params, fetch='one')['n']

    rows = query(f"""
        SELECT al.id, al.action, al.detail, al.ip_address, al.created_at,
               al.user_id, u.name as user_name, u.email as user_email, u.role as user_role,
               al.school_id, s.name as school_name
        FROM activity_logs al
        LEFT JOIN users u ON u.id=al.user_id
        LEFT JOIN schools s ON s.id=al.school_id
        WHERE {where}
        ORDER BY al.created_at DESC
        LIMIT %s OFFSET %s
    """, params + [page_size, offset])

    actions = query("SELECT DISTINCT action FROM activity_logs ORDER BY action")

    return jsonify({
        'logs': [dict(r) for r in rows],
        'total': total,
        'page': page,
        'page_size': page_size,
        'actions': [r['action'] for r in actions],
    })


@super_admin_bp.route('/api/super-admin/activity-logs/export', methods=['GET'])
@require_super_admin
def export_activity_logs():
    """Export hasil filter yang sama persis ke Excel — dibatasi 5000 baris
    terbaru supaya tidak membebani memori/permintaan tunggal."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    where, params = _activity_log_filters()
    rows = query(f"""
        SELECT al.created_at, s.name as school_name, u.name as user_name,
               u.email as user_email, u.role as user_role, al.action, al.detail, al.ip_address
        FROM activity_logs al
        LEFT JOIN users u ON u.id=al.user_id
        LEFT JOIN schools s ON s.id=al.school_id
        WHERE {where}
        ORDER BY al.created_at DESC
        LIMIT 5000
    """, params)

    wb = Workbook(); ws = wb.active
    ws.title = 'Log Aktivitas'
    header = ['Waktu', 'Sekolah', 'Nama User', 'Email', 'Peran', 'Aksi', 'Detail', 'IP Address']
    ws.append(header)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='0F4C35')
        cell.alignment = Alignment(horizontal='center')

    for r in rows:
        ws.append([
            str(r['created_at'])[:19] if r['created_at'] else '',
            r['school_name'] or '',
            r['user_name'] or '(terhapus)',
            r['user_email'] or '',
            r['user_role'] or '',
            r['action'] or '',
            r['detail'] or '',
            r['ip_address'] or '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(60, max(len(str(c.value or '')) for c in col) + 4)

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='log_aktivitas.xlsx')
