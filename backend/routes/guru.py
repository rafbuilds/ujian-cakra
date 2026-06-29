# backend/routes/guru.py
from flask import Blueprint, request, jsonify
import uuid
from db import query
from auth import require_guru, require_auth

guru_bp = Blueprint('guru', __name__)

# ── Taught Classes ─────────────────────────────────────────────
@guru_bp.route('/api/guru/taught-classes', methods=['GET'])
@require_guru
def get_taught_classes():
    rows = query("""
        SELECT c.* FROM classes c
        JOIN guru_classes gc ON gc.class_id=c.id
        WHERE gc.teacher_id=%s AND c.school_id=%s
        ORDER BY c.name
    """, (request.user_id, request.school_id))
    return jsonify([dict(r) for r in rows])

@guru_bp.route('/api/guru/taught-classes', methods=['POST'])
@require_guru
def add_taught_class():
    class_id = (request.json or {}).get('class_id','').strip()
    if not class_id: return jsonify({'error': 'class_id wajib'}), 400
    cls = query("SELECT id FROM classes WHERE id=%s AND school_id=%s", (class_id, request.school_id), fetch='one')
    if not cls: return jsonify({'error': 'Kelas tidak ditemukan'}), 404
    query("INSERT INTO guru_classes (teacher_id, class_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
          (request.user_id, class_id), fetch='none')
    return jsonify({'ok': True})

@guru_bp.route('/api/guru/taught-classes/<class_id>', methods=['DELETE'])
@require_guru
def remove_taught_class(class_id):
    query("DELETE FROM guru_classes WHERE teacher_id=%s AND class_id=%s",
          (request.user_id, class_id), fetch='none')
    return jsonify({'ok': True})

# ── Subjects (Mapel per guru) ──────────────────────────────────
@guru_bp.route('/api/subjects', methods=['GET'])
@require_guru
def get_subjects():
    if getattr(request, 'user_role', '') in ('admin', 'super_admin'):
        rows = query("SELECT * FROM subjects WHERE school_id=%s ORDER BY name", (request.school_id,))
    else:
        rows = query(
            "SELECT * FROM subjects WHERE teacher_id=%s AND school_id=%s ORDER BY name",
            (request.user_id, request.school_id)
        )
    return jsonify([dict(r) for r in rows])

@guru_bp.route('/api/subjects', methods=['POST'])
@require_guru
def create_subject():
    name = (request.json or {}).get('name','').strip()
    if not name: return jsonify({'error': 'Nama mapel wajib'}), 400
    existing = query("SELECT id FROM subjects WHERE LOWER(name)=LOWER(%s) AND teacher_id=%s AND school_id=%s",
                     (name, request.user_id, request.school_id), fetch='one')
    if existing: return jsonify({'error': 'Mapel sudah ada'}), 409
    sub = query("INSERT INTO subjects (id, name, teacher_id, school_id) VALUES (%s,%s,%s,%s) RETURNING *",
                (str(uuid.uuid4()), name, request.user_id, request.school_id), fetch='one')
    return jsonify(dict(sub)), 201

@guru_bp.route('/api/subjects/<subject_id>', methods=['PATCH'])
@require_guru
def update_subject(subject_id):
    name = (request.json or {}).get('name','').strip()
    if not name: return jsonify({'error': 'Nama wajib'}), 400
    query("UPDATE subjects SET name=%s WHERE id=%s AND teacher_id=%s AND school_id=%s",
          (name, subject_id, request.user_id, request.school_id), fetch='none')
    return jsonify(dict(query("SELECT * FROM subjects WHERE id=%s AND school_id=%s",
                              (subject_id, request.school_id), fetch='one')))

@guru_bp.route('/api/subjects/<subject_id>', methods=['DELETE'])
@require_guru
def delete_subject(subject_id):
    sub = query("SELECT * FROM subjects WHERE id=%s AND school_id=%s", (subject_id, request.school_id), fetch='one')
    if not sub: return jsonify({'error': 'Tidak ditemukan'}), 404
    if str(sub.get('teacher_id','')) != request.user_id and request.user_role not in ('admin', 'super_admin'):
        return jsonify({'error': 'Bukan mapel kamu'}), 403
    used = query("SELECT id FROM exams WHERE subject_id=%s LIMIT 1", (subject_id,), fetch='one')
    if used: return jsonify({'error': 'Mapel masih dipakai di ujian'}), 400
    query("DELETE FROM subjects WHERE id=%s AND school_id=%s", (subject_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@guru_bp.route('/api/mapel-referensi', methods=['GET'])
@require_guru
def mapel_referensi():
    rows = query("SELECT DISTINCT name FROM subjects WHERE school_id=%s ORDER BY name", (request.school_id,))
    return jsonify([r['name'] for r in rows])

# ── Siswa (guru view) ──────────────────────────────────────────
# FIX: return {students: [...]} bukan raw array, sesuai yang dibutuhkan frontend
@guru_bp.route('/api/guru/siswa', methods=['GET'])
@require_guru
def get_guru_siswa():
    # Cek apakah guru sudah assign kelas
    taught = query("SELECT class_id FROM guru_classes WHERE teacher_id=%s", (request.user_id,))

    if taught:
        # Guru sudah set kelas — tampilkan siswa dari kelas yang diajar saja
        rows = query("""
            SELECT
                u.id, u.name, u.email, u.nisn, u.class_id,
                c.name as class_name,
                ROUND(AVG(r.score)::numeric, 1) as avg_score
            FROM users u
            LEFT JOIN classes c ON c.id=u.class_id
            LEFT JOIN exam_sessions es ON es.student_id=u.id AND es.submitted_at IS NOT NULL
            LEFT JOIN results r ON r.session_id=es.id
            WHERE u.role='siswa' AND u.school_id=%s AND u.class_id IN (
                SELECT class_id FROM guru_classes WHERE teacher_id=%s
            )
            GROUP BY u.id, u.name, u.email, u.nisn, u.class_id, c.name
            ORDER BY c.name, u.name
        """, (request.school_id, request.user_id))
    else:
        # Guru belum set kelas — tampilkan semua siswa
        rows = query("""
            SELECT
                u.id, u.name, u.email, u.nisn, u.class_id,
                c.name as class_name,
                ROUND(AVG(r.score)::numeric, 1) as avg_score
            FROM users u
            LEFT JOIN classes c ON c.id=u.class_id
            LEFT JOIN exam_sessions es ON es.student_id=u.id AND es.submitted_at IS NOT NULL
            LEFT JOIN results r ON r.session_id=es.id
            WHERE u.role='siswa' AND u.school_id=%s
            GROUP BY u.id, u.name, u.email, u.nisn, u.class_id, c.name
            ORDER BY c.name, u.name
        """, (request.school_id,))

    students = [dict(r) for r in rows]
    return jsonify({'students': students, 'total': len(students)})

@guru_bp.route('/api/guru/siswa/import', methods=['POST'])
@require_guru
def guru_import_siswa():
    from openpyxl import load_workbook
    file = request.files.get('file')
    if not file: return jsonify({'error': 'File tidak ada'}), 400
    wb = load_workbook(file); ws = wb.active
    saved = 0
    errors = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]: continue
        try:
            name = str(row[0] or '').strip()
            nisn = str(row[1] or '').strip()
            class_id = str(row[2] or '').strip() or None
            email = str(row[3] or '').strip() or None
            if not name: continue

            # Validasi kelas kalau ada
            if class_id:
                cls = query("SELECT id FROM classes WHERE id=%s AND school_id=%s",
                           (class_id, request.school_id), fetch='one')
                if not cls:
                    errors.append(f"Baris {idx}: Kelas '{class_id}' tidak ditemukan")
                    class_id = None

            existing = query("SELECT id FROM users WHERE email=%s AND school_id=%s",
                             (email, request.school_id), fetch='one') if email else None
            if existing:
                query("UPDATE users SET name=%s, nisn=%s, class_id=%s WHERE id=%s",
                      (name, nisn or None, class_id, existing['id']), fetch='none')
            else:
                dummy_gid = f"import_{uuid.uuid4().hex[:12]}"
                query("""INSERT INTO users (id, google_id, email, name, nisn, class_id, role, is_active, school_id)
                         VALUES (%s,%s,%s,%s,%s,%s,'siswa',true,%s) ON CONFLICT (google_id) DO NOTHING""",
                      (str(uuid.uuid4()), dummy_gid, email, name, nisn or None, class_id, request.school_id), fetch='none')
            saved += 1
        except Exception as e:
            errors.append(f"Baris {idx}: {str(e)}")
    return jsonify({'ok': True, 'saved': saved, 'errors': errors})

# ── Riwayat ujian satu siswa ───────────────────────────────────
@guru_bp.route('/api/siswa/<student_id>/history', methods=['GET'])
@require_guru
def siswa_history(student_id):
    rows = query("""
        SELECT
            es.id, es.submitted_at, es.started_at,
            e.title as exam_title, e.id as exam_id,
            e.duration_minutes,
            s.name as subject_name,
            r.score, r.correct_count, r.wrong_count, r.empty_count,
            es.tab_violations
        FROM exam_sessions es
        JOIN exams e ON e.id=es.exam_id
        LEFT JOIN subjects s ON s.id=e.subject_id
        LEFT JOIN results r ON r.session_id=es.id
        WHERE es.student_id=%s AND e.school_id=%s
        ORDER BY COALESCE(es.submitted_at, es.started_at) DESC
    """, (student_id, request.school_id))
    return jsonify({'history': [dict(r) for r in rows]})

# ── Classes (shared endpoint) ──────────────────────────────────
@guru_bp.route('/api/classes', methods=['GET'])
@require_auth
def get_classes():
    rows = query("SELECT * FROM classes WHERE school_id=%s ORDER BY name", (request.school_id,))
    return jsonify([dict(r) for r in rows])

@guru_bp.route('/api/template-siswa', methods=['GET'])
@require_guru
def template_siswa():
    from flask import send_file
    from openpyxl import Workbook
    import io
    wb = Workbook(); ws = wb.active; ws.title = 'Data Siswa'
    ws.append(['Nama Lengkap','NISN','ID Kelas','Email'])
    ws.append(['Adi Saputra','0012345678','x_1','adi@sman1batangan.sch.id'])
    ws.append(['Budi Santoso','0098765432','x_2','budi@sman1batangan.sch.id'])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='template_siswa.xlsx')

# ── Sessions (allow exit) ──────────────────────────────────────
@guru_bp.route('/api/sessions/<session_id>/allow-exit', methods=['POST'])
@require_guru
def allow_exit(session_id):
    query("""UPDATE exam_sessions SET exit_allowed=true
              WHERE id=%s AND school_id=%s""", (session_id, request.school_id), fetch='none')
    return jsonify({'ok': True})
# ── Tahun Ajaran / Semester (read-only, untuk navigasi Room Ujian) ──
@guru_bp.route('/api/guru/academic-years', methods=['GET'])
@require_guru
def guru_list_academic_years():
    rows = query("SELECT id, name, is_active FROM academic_years WHERE school_id=%s ORDER BY name DESC",
                 (request.school_id,))
    return jsonify([dict(r) for r in rows])

@guru_bp.route('/api/guru/semesters', methods=['GET'])
@require_guru
def guru_list_semesters():
    year_id = request.args.get('academic_year_id')
    if year_id:
        rows = query("""SELECT id, name, is_active, academic_year_id FROM semesters
                        WHERE academic_year_id=%s AND school_id=%s ORDER BY name""",
                     (year_id, request.school_id))
    else:
        rows = query("SELECT id, name, is_active, academic_year_id FROM semesters WHERE school_id=%s ORDER BY name",
                     (request.school_id,))
    return jsonify([dict(r) for r in rows])

@guru_bp.route('/api/guru/active-period', methods=['GET'])
@require_guru
def guru_active_period():
    year = query("SELECT id, name FROM academic_years WHERE is_active=true AND school_id=%s LIMIT 1",
                 (request.school_id,), fetch='one')
    sem  = query("SELECT id, name, academic_year_id FROM semesters WHERE is_active=true AND school_id=%s LIMIT 1",
                 (request.school_id,), fetch='one')
    return jsonify({
        'academic_year': dict(year) if year else None,
        'semester': dict(sem) if sem else None,
    })
