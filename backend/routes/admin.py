# backend/routes/admin.py
from flask import Blueprint, request, jsonify, send_file
import uuid, io
from db import query, log_activity
from auth import require_admin, require_guru

admin_bp = Blueprint('admin', __name__)

# ── Users ──────────────────────────────────────────────────────
@admin_bp.route('/api/admin/users', methods=['GET'])
@require_admin
def get_users():
    try:
        rows = query("""
            SELECT id, email, name, role, avatar_url, last_login, device_id, created_at,
                   (password_hash IS NOT NULL AND password_hash != '') AS has_password
            FROM users WHERE school_id=%s ORDER BY role, name
        """, (request.school_id,))
    except Exception:
        # Fallback jika kolom password_hash belum ada (migration belum dijalankan)
        rows = query("""
            SELECT id, email, name, role, avatar_url, last_login, device_id, created_at
            FROM users WHERE school_id=%s ORDER BY role, name
        """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/admin/users/<user_id>', methods=['PATCH'])
@require_admin
def update_user(user_id):
    data = request.json or {}
    allowed = ['role', 'name', 'email']
    for f in [k for k in data if k in allowed]:
        query(f"UPDATE users SET {f}=%s WHERE id=%s AND school_id=%s",
              (data[f], user_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/users/<user_id>/set-password', methods=['POST'])
@require_admin
def set_user_password(user_id):
    from auth import hash_password
    pw = (request.json or {}).get('password', '').strip()
    if not pw or len(pw) < 6:
        return jsonify({'error': 'Password minimal 6 karakter'}), 400
    try:
        query("UPDATE users SET password_hash=%s WHERE id=%s AND school_id=%s",
              (hash_password(pw), user_id, request.school_id), fetch='none')
        return jsonify({'ok': True})
    except Exception as e:
        if 'password_hash' in str(e):
            return jsonify({'error': 'Kolom password_hash belum ada. Jalankan dulu di Supabase SQL Editor:\nALTER TABLE users ADD COLUMN IF NOT EXISTS password_hash TEXT;'}), 500
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/api/admin/users', methods=['POST'])
@require_admin
def create_user():
    from auth import hash_password, validate_email_domain
    data  = request.json or {}
    name  = data.get('name', '').strip()
    email = data.get('email', '').strip().lower()
    role  = data.get('role', 'guru')
    pw    = data.get('password', '').strip()
    if not name or not email:
        return jsonify({'error': 'Nama dan email wajib'}), 400
    if not pw or len(pw) < 6:
        return jsonify({'error': 'Password minimal 6 karakter'}), 400
    domain_err = validate_email_domain(request.school_id, email)
    if domain_err:
        return jsonify({'error': domain_err}), 400
    existing = query("SELECT id FROM users WHERE LOWER(email)=%s", (email,), fetch='one')
    if existing:
        return jsonify({'error': 'Email sudah terdaftar'}), 409
    dummy_gid = f"manual_{uuid.uuid4().hex[:12]}"
    uid = str(uuid.uuid4())
    try:
        query("""INSERT INTO users (id, google_id, email, name, role, password_hash, is_active, school_id)
                 VALUES (%s,%s,%s,%s,%s,%s,true,%s)""",
              (uid, dummy_gid, email, name, role, hash_password(pw), request.school_id), fetch='none')
    except Exception as e:
        err = str(e)
        if 'password_hash' in err:
            return jsonify({'error': 'Kolom password_hash belum ada. Jalankan migration SQL di Supabase dulu.'}), 500
        if 'google_id' in err and 'null' in err.lower():
            return jsonify({'error': 'Kolom google_id masih NOT NULL. Jalankan: ALTER TABLE users ALTER COLUMN google_id DROP NOT NULL'}), 500
        return jsonify({'error': 'Gagal: ' + err}), 500
    user = query("SELECT id,email,name,role FROM users WHERE id=%s", (uid,), fetch='one')
    return jsonify(dict(user)), 201

@admin_bp.route('/api/admin/users/<user_id>', methods=['DELETE'])
@require_admin
def delete_user(user_id):
    query("DELETE FROM users WHERE id=%s AND role!='admin' AND school_id=%s",
          (user_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

# ── Siswa ──────────────────────────────────────────────────────
@admin_bp.route('/api/admin/siswa', methods=['GET'])
@require_admin
def get_siswa():
    grade    = request.args.get('grade', '')
    class_id = request.args.get('class_id', '')
    search   = request.args.get('search', '')
    page     = int(request.args.get('page', 1))
    # Cap dinaikkan dari 200 ke 5000 — sekolah dengan >200 siswa sebelumnya
    # kepotong diam-diam di halaman Data Siswa (cuma 200 pertama kebaca),
    # padahal dashboard admin sudah benar hitung COUNT(*) semua siswa.
    per_page = min(int(request.args.get('per_page', 50)), 5000)

    where = ["u.role='siswa'", "u.school_id=%s"]
    params = [request.school_id]
    if grade:    where.append("c.grade=%s"); params.append(int(grade))
    if class_id: where.append("u.class_id=%s"); params.append(class_id)
    if search:   where.append("u.name ILIKE %s"); params.append(f'%{search}%')

    where_sql = ' AND '.join(where)
    total = query(f"""
        SELECT COUNT(*) as n FROM users u
        LEFT JOIN classes c ON c.id=u.class_id
        WHERE {where_sql}
    """, params, fetch='one')['n']

    rows = query(f"""
        SELECT u.id, u.name, u.email, u.nisn, u.class_id,
               c.name as class_name, c.grade,
               u.device_id, u.device_info, u.last_login, u.is_active
        FROM users u
        LEFT JOIN classes c ON c.id=u.class_id
        WHERE {where_sql}
        ORDER BY c.grade, LENGTH(c.id), c.id, u.name
        LIMIT %s OFFSET %s
    """, params + [per_page, (page-1)*per_page])

    return jsonify({'data': [dict(r) for r in rows], 'total': total, 'page': page})

@admin_bp.route('/api/admin/siswa', methods=['POST'])
@require_admin
def create_siswa():
    from auth import validate_email_domain
    data     = request.json or {}
    name     = data.get('name', '').strip()
    email    = (data.get('email') or '').strip().lower()
    nisn     = (data.get('nisn') or '').strip()
    class_id = data.get('class_id') or None
    if not name:
        return jsonify({'error': 'Nama wajib diisi'}), 400
    if email:
        domain_err = validate_email_domain(request.school_id, email)
        if domain_err:
            return jsonify({'error': domain_err}), 400
        existing = query("SELECT id FROM users WHERE LOWER(email)=%s", (email,), fetch='one')
        if existing:
            return jsonify({'error': 'Email sudah dipakai akun lain'}), 409
    dummy_gid = f"manual_{uuid.uuid4().hex[:12]}"
    uid = str(uuid.uuid4())
    query("""INSERT INTO users (id, google_id, email, name, nisn, class_id, role, is_active, school_id)
             VALUES (%s,%s,%s,%s,%s,%s,'siswa',true,%s)""",
          (uid, dummy_gid, email or None, name, nisn or None, class_id, request.school_id), fetch='none')
    user = query("""
        SELECT u.id, u.name, u.email, u.nisn, u.class_id, c.name as class_name, c.grade,
               u.device_id, u.device_info, u.last_login, u.is_active
        FROM users u LEFT JOIN classes c ON c.id=u.class_id WHERE u.id=%s
    """, (uid,), fetch='one')
    return jsonify(dict(user)), 201

@admin_bp.route('/api/admin/siswa/<siswa_id>', methods=['PATCH'])
@require_admin
def update_siswa(siswa_id):
    from auth import validate_email_domain
    data = request.json or {}
    if 'email' in data:
        email = (data['email'] or '').strip().lower()
        if email:
            domain_err = validate_email_domain(request.school_id, email)
            if domain_err:
                return jsonify({'error': domain_err}), 400
            existing = query("SELECT id FROM users WHERE LOWER(email)=%s AND id!=%s",
                             (email, siswa_id), fetch='one')
            if existing:
                return jsonify({'error': 'Email sudah dipakai akun lain'}), 409
        query("UPDATE users SET email=%s WHERE id=%s AND role='siswa' AND school_id=%s",
              (email or None, siswa_id, request.school_id), fetch='none')
    allowed = ['name', 'nisn', 'class_id', 'is_active']
    for f in [k for k in data if k in allowed]:
        query(f"UPDATE users SET {f}=%s WHERE id=%s AND role='siswa' AND school_id=%s",
              (data[f] or None, siswa_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/siswa/<siswa_id>', methods=['DELETE'])
@require_admin
def delete_siswa(siswa_id):
    query("DELETE FROM users WHERE id=%s AND role='siswa' AND school_id=%s",
          (siswa_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/siswa/<siswa_id>/reset-device', methods=['POST'])
@require_admin
def reset_siswa_device(siswa_id):
    query("UPDATE users SET device_id=NULL, device_info=NULL WHERE id=%s AND role='siswa' AND school_id=%s",
          (siswa_id, request.school_id), fetch='none')
    return jsonify({'ok': True, 'message': 'Device berhasil direset'})

def _upsert_siswa_row(name, nisn, class_id, email, school_id):
    """Helper: insert atau update satu baris data siswa."""
    if not name: return False
    from auth import validate_email_domain
    if email and validate_email_domain(school_id, email):
        return False  # domain email tidak cocok dengan yang diset super_admin
    dummy_gid = f"import_{uuid.uuid4().hex[:12]}"
    existing = query("SELECT id FROM users WHERE email=%s AND role='siswa' AND school_id=%s",
                     (email, school_id), fetch='one') if email else None
    if existing:
        query("UPDATE users SET name=%s, nisn=%s, class_id=%s WHERE id=%s",
              (name, nisn or None, class_id or None, existing['id']), fetch='none')
    else:
        query("""INSERT INTO users (id, google_id, email, name, nisn, class_id, role, is_active, school_id)
                 VALUES (%s,%s,%s,%s,%s,%s,'siswa',true,%s)
                 ON CONFLICT (google_id) DO NOTHING""",
              (str(uuid.uuid4()), dummy_gid, email or None, name,
               nisn or None, class_id or None, school_id), fetch='none')
    return True

@admin_bp.route('/api/admin/siswa/import', methods=['POST'])
@require_admin
def import_siswa():
    file = request.files.get('file')
    if not file: return jsonify({'error': 'File tidak ada'}), 400
    fname = (file.filename or '').lower()
    imported = 0
    school_id = request.school_id

    if fname.endswith('.docx'):
        from docx import Document
        doc = Document(file)
        for table in doc.tables:
            for i, row in enumerate(table.rows):
                if i == 0: continue  # skip header
                cells = [c.text.strip() for c in row.cells]
                if len(cells) < 1 or not cells[0]: continue
                name     = cells[0] if len(cells) > 0 else ''
                nisn     = cells[1] if len(cells) > 1 else ''
                class_id = cells[2] if len(cells) > 2 else ''
                email    = cells[3] if len(cells) > 3 else ''
                if _upsert_siswa_row(name, nisn, class_id, email, school_id):
                    imported += 1
    else:
        from openpyxl import load_workbook
        wb = load_workbook(file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]: continue
            name, nisn, class_id, email = (str(row[i] or '').strip() for i in range(4))
            if _upsert_siswa_row(name, nisn, class_id, email, school_id):
                imported += 1

    return jsonify({'ok': True, 'imported': imported})

@admin_bp.route('/api/admin/siswa/template', methods=['GET'])
@require_admin
def siswa_template():
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    ws.title = 'Data Siswa'
    ws.append(['Nama Lengkap','NISN','ID Kelas','Email'])
    ws.append(['Contoh: Adi Saputra','0012345678','x_1','adi@sman1batangan.sch.id'])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='template_siswa.xlsx')

@admin_bp.route('/api/admin/siswa/export', methods=['GET'])
@require_admin
def export_siswa():
    from auth import feature_blocked_reason
    blocked = feature_blocked_reason(request.school_id, 'export')
    if blocked:
        return jsonify({'error': blocked}), 403
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    grade    = request.args.get('grade', '')
    class_id = request.args.get('class_id', '')

    where  = ["u.role='siswa'", "u.school_id=%s"]
    params = [request.school_id]
    if grade:    where.append("c.grade=%s");    params.append(int(grade))
    if class_id: where.append("u.class_id=%s"); params.append(class_id)

    rows = query(f"""
        SELECT u.name, u.nisn, u.class_id, c.name as class_name, c.grade,
               u.email, u.device_id, u.last_login, u.is_active
        FROM users u
        LEFT JOIN classes c ON c.id=u.class_id
        WHERE {' AND '.join(where)}
        ORDER BY c.grade, LENGTH(c.id), c.id, u.name
    """, params)

    wb = Workbook(); ws = wb.active
    ws.title = 'Data Siswa'

    header = ['No','Nama Lengkap','NISN','Kelas','Tingkat','Email','Device','Login Terakhir','Status']
    ws.append(header)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color='FFFFFF')
        cell.fill      = PatternFill('solid', fgColor='0F4C35')
        cell.alignment = Alignment(horizontal='center')

    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r['name'] or '',
            r['nisn']  or '',
            r['class_name'] or '',
            r['grade']  or '',
            r['email']  or '',
            'Terdaftar' if r['device_id'] else 'Belum',
            str(r['last_login'])[:16] if r['last_login'] else '',
            'Aktif' if r['is_active'] else 'Nonaktif',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(c.value or '')) for c in col) + 4

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    label = f"_kelas{grade}" if grade else (f"_{class_id}" if class_id else "_semua")
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=f'data_siswa{label}.xlsx')

# ── Classes ────────────────────────────────────────────────────
@admin_bp.route('/api/admin/classes-detail', methods=['GET'])
@require_admin
def classes_detail():
    rows = query("""
        SELECT c.*, COUNT(u.id) as student_count
        FROM classes c LEFT JOIN users u ON u.class_id=c.id AND u.role='siswa'
        WHERE c.school_id=%s
        GROUP BY c.id ORDER BY c.grade, LENGTH(c.id), c.id
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

# ── Subjects ───────────────────────────────────────────────────
@admin_bp.route('/api/admin/subjects', methods=['GET'])
@require_admin
def admin_get_subjects():
    rows = query("""
        SELECT s.*, u.name as teacher_name FROM subjects s
        LEFT JOIN users u ON u.id=s.teacher_id
        WHERE s.school_id=%s
        ORDER BY u.name, s.name
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

# ── Guru Subjects & Classes ────────────────────────────────────
@admin_bp.route('/api/admin/guru/<guru_id>/subjects', methods=['GET'])
@require_admin
def admin_guru_subjects(guru_id):
    rows = query("SELECT * FROM subjects WHERE teacher_id=%s AND school_id=%s ORDER BY name",
                 (guru_id, request.school_id))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/admin/guru/<guru_id>/subjects', methods=['POST'])
@require_admin
def admin_add_guru_subject(guru_id):
    name = (request.json or {}).get('name','').strip()
    if not name: return jsonify({'error': 'Nama mapel wajib'}), 400
    existing = query("SELECT id FROM subjects WHERE LOWER(name)=LOWER(%s) AND teacher_id=%s AND school_id=%s",
                     (name, guru_id, request.school_id), fetch='one')
    if existing: return jsonify({'error': 'Mapel sudah ada'}), 409
    sub = query("INSERT INTO subjects (id, name, teacher_id, school_id) VALUES (%s,%s,%s,%s) RETURNING *",
                (str(uuid.uuid4()), name, guru_id, request.school_id), fetch='one')
    return jsonify(dict(sub)), 201

@admin_bp.route('/api/admin/guru/<guru_id>/subjects/<subject_id>', methods=['DELETE'])
@require_admin
def admin_del_guru_subject(guru_id, subject_id):
    used = query("SELECT id FROM exams WHERE subject_id=%s LIMIT 1", (subject_id,), fetch='one')
    if used: return jsonify({'error': 'Mapel masih dipakai di ujian'}), 400
    query("DELETE FROM subjects WHERE id=%s AND teacher_id=%s AND school_id=%s",
          (subject_id, guru_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/guru/<guru_id>/classes', methods=['GET'])
@require_admin
def admin_guru_classes(guru_id):
    rows = query("""
        SELECT c.* FROM classes c JOIN guru_classes gc ON gc.class_id=c.id
        WHERE gc.teacher_id=%s AND c.school_id=%s ORDER BY c.grade, LENGTH(c.id), c.id
    """, (guru_id, request.school_id))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/admin/guru/<guru_id>/classes', methods=['POST'])
@require_admin
def admin_add_guru_class(guru_id):
    class_id = (request.json or {}).get('class_id','').strip()
    if not class_id: return jsonify({'error': 'class_id wajib'}), 400
    query("INSERT INTO guru_classes (teacher_id, class_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
          (guru_id, class_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/guru/<guru_id>/classes/<class_id>', methods=['DELETE'])
@require_admin
def admin_del_guru_class(guru_id, class_id):
    query("DELETE FROM guru_classes WHERE teacher_id=%s AND class_id=%s",
          (guru_id, class_id), fetch='none')
    return jsonify({'ok': True})

# ── Exams (admin view) ─────────────────────────────────────────
@admin_bp.route('/api/admin/questions/check-similar', methods=['POST'])
@require_admin
def admin_check_similar_questions():
    """Admin bisa cek soal serupa lintas semua guru (admin sudah punya akses
    penuh ke semua soal)."""
    content = (request.json or {}).get('content', '').strip()
    if len(content) < 10:
        return jsonify([])
    try:
        rows = query("""
            SELECT q.id as question_id, q.content, q.exam_id, e.title as exam_title,
                   u.name as teacher_name, similarity(q.content, %s) as score
            FROM questions q
            JOIN exams e ON e.id = q.exam_id
            JOIN users u ON u.id = e.teacher_id
            WHERE similarity(q.content, %s) > 0.3 AND e.school_id=%s
            ORDER BY score DESC LIMIT 10
        """, (content, content, request.school_id))
        return jsonify([dict(r) for r in rows])
    except Exception:
        return jsonify([])

@admin_bp.route('/api/admin/exams', methods=['GET'])
@require_admin
def admin_get_exams():
    rows = query("""
        SELECT e.*, u.name as teacher_name, s.name as subject_name, r.name as room_name,
               sem.name as semester_name, ay.name as academic_year_name,
               (SELECT COUNT(*) FROM questions q WHERE q.exam_id=e.id) as question_count,
               (SELECT STRING_AGG(c.name, ', ') FROM exam_classes ec JOIN classes c ON c.id=ec.class_id WHERE ec.exam_id=e.id) as class_names
        FROM exams e
        LEFT JOIN users u ON u.id=e.teacher_id
        LEFT JOIN subjects s ON s.id=e.subject_id
        LEFT JOIN rooms r ON r.id=e.room_id
        LEFT JOIN semesters sem ON sem.id=e.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE e.school_id=%s
        ORDER BY e.created_at DESC
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/admin/exams/<exam_id>', methods=['DELETE'])
@require_admin
def admin_delete_exam(exam_id):
    query("DELETE FROM exams WHERE id=%s AND school_id=%s", (exam_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/exams/<exam_id>/force-finish', methods=['POST'])
@require_admin
def admin_force_finish(exam_id):
    exam_data = query("SELECT * FROM exams WHERE id=%s AND school_id=%s",
                      (exam_id, request.school_id), fetch='one')
    if not exam_data:
        return jsonify({'error': 'Ujian tidak ditemukan'}), 404
    ongoing = query("""
        SELECT id FROM exam_sessions WHERE exam_id=%s AND submitted_at IS NULL
    """, (exam_id,))
    total_q = query("SELECT COUNT(*) as n FROM questions WHERE exam_id=%s", (exam_id,), fetch='one')['n']
    spc = float(exam_data.get('score_per_correct') or (100.0/total_q if total_q else 0))

    for sess in ongoing:
        sid = str(sess['id'])
        correct = query("""SELECT COUNT(*) as n FROM answers a JOIN options o ON o.id=a.option_id
                           WHERE a.session_id=%s AND o.is_correct=true""", (sid,), fetch='one')['n']
        wrong   = query("""SELECT COUNT(*) as n FROM answers a JOIN options o ON o.id=a.option_id
                           WHERE a.session_id=%s AND o.is_correct=false""", (sid,), fetch='one')['n']
        empty   = total_q - correct - wrong
        score   = round(correct * spc, 2)
        query("UPDATE exam_sessions SET submitted_at=NOW(), auto_submitted=true WHERE id=%s", (sid,), fetch='none')
        ex = query("SELECT id FROM results WHERE session_id=%s", (sid,), fetch='one')
        if ex:
            query("UPDATE results SET score=%s,correct_count=%s,wrong_count=%s,empty_count=%s WHERE session_id=%s",
                  (score, correct, wrong, empty, sid), fetch='none')
        else:
            query("INSERT INTO results (id,session_id,score,correct_count,wrong_count,empty_count) VALUES (%s,%s,%s,%s,%s,%s)",
                  (str(uuid.uuid4()), sid, score, correct, wrong, empty), fetch='none')

    query("UPDATE exams SET status='finished' WHERE id=%s", (exam_id,), fetch='none')
    return jsonify({'ok': True, 'submitted': len(ongoing)})

# ── Exam Settings ──────────────────────────────────────────────
@admin_bp.route('/api/admin/exam-settings', methods=['GET'])
@require_admin
def get_exam_settings():
    s = query("SELECT * FROM exam_settings WHERE school_id=%s LIMIT 1", (request.school_id,), fetch='one')
    return jsonify(dict(s) if s else {
        'passing_grade': 75, 'allow_remedial': True,
        'max_violations': 5, 'auto_submit_on_violation': True, 'show_ranking': True,
        'proctor_code': None
    })

@admin_bp.route('/api/admin/exam-settings', methods=['PATCH'])
@require_admin
def update_exam_settings():
    _ALLOWED = {'passing_grade', 'allow_remedial', 'max_violations',
                'auto_submit_on_violation', 'show_ranking', 'proctor_code'}
    data = {k: v for k, v in (request.json or {}).items() if k in _ALLOWED}
    if not data:
        return jsonify({'error': 'Tidak ada field yang valid'}), 400
    existing = query("SELECT id FROM exam_settings WHERE school_id=%s LIMIT 1",
                     (request.school_id,), fetch='one')
    if existing:
        fields = ', '.join(f"{k}=%s" for k in data)
        query(f"UPDATE exam_settings SET {fields} WHERE id=%s",
              list(data.values()) + [str(existing['id'])], fetch='none')
    else:
        cols = ', '.join(list(data.keys()) + ['school_id'])
        vals = ', '.join(['%s'] * (len(data) + 1))
        query(f"INSERT INTO exam_settings ({cols}) VALUES ({vals})",
              list(data.values()) + [request.school_id], fetch='none')
    return jsonify({'ok': True})

# ── Rooms ──────────────────────────────────────────────────────
@admin_bp.route('/api/admin/rooms', methods=['GET'])
@require_admin
def get_rooms():
    rows = query("""
        SELECT r.*,
               u.name as created_by_name,
               sem.name as semester_name, ay.name as academic_year_name,
               COUNT(DISTINCT rt.teacher_id) as teacher_count,
               COUNT(DISTINCT rc.class_id) as class_count
        FROM rooms r
        LEFT JOIN users u ON u.id=r.created_by
        LEFT JOIN room_teachers rt ON rt.room_id=r.id
        LEFT JOIN room_classes rc ON rc.room_id=r.id
        LEFT JOIN semesters sem ON sem.id=r.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE r.school_id=%s
        GROUP BY r.id, u.name, sem.name, ay.name
        ORDER BY r.created_at DESC
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/admin/rooms', methods=['POST'])
@require_admin
def create_room():
    data = request.json or {}
    name = data.get('name','').strip()
    if not name: return jsonify({'error': 'Nama room wajib'}), 400
    room = query("""
        INSERT INTO rooms (id, name, description, created_by, semester_id, school_id)
        VALUES (%s,%s,%s,%s,%s,%s) RETURNING *
    """, (str(uuid.uuid4()), name, data.get('description',''), request.user_id,
          data.get('semester_id') or None, request.school_id), fetch='one')
    # Assign classes
    for cls in (data.get('class_ids') or []):
        query("INSERT INTO room_classes (room_id,class_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
              (str(room['id']), cls), fetch='none')
    # Assign teachers
    for tid in (data.get('teacher_ids') or []):
        query("INSERT INTO room_teachers (room_id,teacher_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
              (str(room['id']), tid), fetch='none')
    return jsonify(dict(room)), 201

@admin_bp.route('/api/admin/rooms/<room_id>', methods=['PATCH'])
@require_admin
def update_room(room_id):
    data = request.json or {}
    if 'name' in data:
        query("UPDATE rooms SET name=%s WHERE id=%s AND school_id=%s", (data['name'], room_id, request.school_id), fetch='none')
    if 'description' in data:
        query("UPDATE rooms SET description=%s WHERE id=%s AND school_id=%s", (data['description'], room_id, request.school_id), fetch='none')
    if 'is_active' in data:
        query("UPDATE rooms SET is_active=%s WHERE id=%s AND school_id=%s", (data['is_active'], room_id, request.school_id), fetch='none')
    if 'semester_id' in data:
        query("UPDATE rooms SET semester_id=%s WHERE id=%s AND school_id=%s", (data['semester_id'] or None, room_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/rooms/<room_id>', methods=['DELETE'])
@require_admin
def delete_room(room_id):
    query("DELETE FROM rooms WHERE id=%s AND school_id=%s", (room_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/rooms/<room_id>', methods=['GET'])
@require_admin
def get_room_detail(room_id):
    room = query("""
        SELECT r.*, sem.name as semester_name, ay.name as academic_year_name
        FROM rooms r
        LEFT JOIN semesters sem ON sem.id=r.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE r.id=%s AND r.school_id=%s
    """, (room_id, request.school_id), fetch='one')
    if not room: return jsonify({'error': 'Tidak ditemukan'}), 404
    teachers = query("""
        SELECT u.id, u.name, u.email, u.avatar_url FROM users u
        JOIN room_teachers rt ON rt.teacher_id=u.id
        WHERE rt.room_id=%s ORDER BY u.name
    """, (room_id,))
    classes = query("""
        SELECT c.* FROM classes c
        JOIN room_classes rc ON rc.class_id=c.id
        WHERE rc.room_id=%s ORDER BY c.grade, LENGTH(c.id), c.id
    """, (room_id,))
    # Hanya ujian yang memang sudah diterapkan (room_id=room ini) — soal yang
    # masih di Bank Soal (belum diterapkan, room_id NULL) tidak ikut muncul.
    exams = query("""
        SELECT e.*, u.name as teacher_name, s.name as subject_name,
               sem.name as semester_name, ay.name as academic_year_name,
               (SELECT COUNT(*) FROM questions q WHERE q.exam_id=e.id) as question_count,
               (SELECT STRING_AGG(c.name, ', ' ORDER BY c.name)
                FROM exam_classes ec JOIN classes c ON c.id=ec.class_id
                WHERE ec.exam_id=e.id) as class_names
        FROM exams e
        JOIN users u ON u.id=e.teacher_id
        LEFT JOIN subjects s ON s.id=e.subject_id
        LEFT JOIN semesters sem ON sem.id=e.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE e.room_id=%s
        ORDER BY e.start_at, u.name, e.created_at DESC
    """, (room_id,))
    return jsonify({
        **dict(room),
        'teachers': [dict(t) for t in teachers],
        'classes': [dict(c) for c in classes],
        'exams': [dict(e) for e in exams],
    })

@admin_bp.route('/api/admin/rooms/<room_id>/teachers', methods=['POST'])
@require_admin
def add_room_teacher(room_id):
    if not query("SELECT 1 FROM rooms WHERE id=%s AND school_id=%s", (room_id, request.school_id), fetch='one'):
        return jsonify({'error': 'Room tidak ditemukan'}), 404
    teacher_id = (request.json or {}).get('teacher_id','')
    query("INSERT INTO room_teachers (room_id,teacher_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
          (room_id, teacher_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/rooms/<room_id>/teachers/<teacher_id>', methods=['DELETE'])
@require_admin
def remove_room_teacher(room_id, teacher_id):
    query("""DELETE FROM room_teachers WHERE room_id=%s AND teacher_id=%s
              AND room_id IN (SELECT id FROM rooms WHERE school_id=%s)""",
          (room_id, teacher_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/rooms/<room_id>/classes', methods=['POST'])
@require_admin
def add_room_class(room_id):
    if not query("SELECT 1 FROM rooms WHERE id=%s AND school_id=%s", (room_id, request.school_id), fetch='one'):
        return jsonify({'error': 'Room tidak ditemukan'}), 404
    data = request.json or {}
    class_ids = data.get('class_ids') or ([data['class_id']] if data.get('class_id') else [])
    if not class_ids:
        return jsonify({'error': 'class_id/class_ids wajib'}), 400
    query("""
        INSERT INTO room_classes (room_id, class_id)
        SELECT %s, c.id FROM UNNEST(%s::text[]) AS c(id)
        ON CONFLICT DO NOTHING
    """, (room_id, class_ids), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/admin/rooms/<room_id>/classes/<class_id>', methods=['DELETE'])
@require_admin
def remove_room_class(room_id, class_id):
    query("""DELETE FROM room_classes WHERE room_id=%s AND class_id=%s
              AND room_id IN (SELECT id FROM rooms WHERE school_id=%s)""",
          (room_id, class_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

# ── Guru: lihat rooms yang dia ikuti ──────────────────────────
@admin_bp.route('/api/guru/rooms', methods=['GET'])
@require_guru
def get_guru_rooms():
    sem_id = request.args.get('semester_id')
    extra = " AND r.semester_id=%s" if sem_id else ""
    params = [request.user_id, request.user_id, request.school_id] + ([sem_id] if sem_id else [])
    rows = query(f"""
        SELECT r.*, sem.name as semester_name, ay.name as academic_year_name,
               COUNT(DISTINCT rc.class_id) as class_count,
               COUNT(DISTINCT e.id) as exam_count
        FROM rooms r
        JOIN room_teachers rt ON rt.room_id=r.id AND rt.teacher_id=%s
        LEFT JOIN room_classes rc ON rc.room_id=r.id
        LEFT JOIN exams e ON e.room_id=r.id AND e.teacher_id=%s
        LEFT JOIN semesters sem ON sem.id=r.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE r.is_active=true AND r.school_id=%s{extra}
        GROUP BY r.id, sem.name, ay.name ORDER BY r.created_at DESC
    """, tuple(params))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/guru/rooms/all', methods=['GET'])
@require_guru
def get_all_rooms_for_guru():
    """Semua rooms aktif + flag is_member untuk guru yang login.
    Filter opsional ?semester_id= agar room dari semester lama tidak tercampur
    dengan room semester yang sedang berjalan."""
    sem_id = request.args.get('semester_id')
    extra = " AND r.semester_id=%s" if sem_id else ""
    params = [request.user_id, request.school_id] + ([sem_id] if sem_id else [])
    rows = query(f"""
        SELECT r.*,
               u.name as created_by_name,
               sem.name as semester_name, ay.name as academic_year_name,
               COUNT(DISTINCT rc.class_id) as class_count,
               COUNT(DISTINCT e.id) as exam_count,
               BOOL_OR(rt2.teacher_id = %s) as is_member
        FROM rooms r
        LEFT JOIN users u ON u.id=r.created_by
        LEFT JOIN room_teachers rt2 ON rt2.room_id=r.id
        LEFT JOIN room_classes rc ON rc.room_id=r.id
        LEFT JOIN exams e ON e.room_id=r.id
        LEFT JOIN semesters sem ON sem.id=r.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE r.is_active=true AND r.school_id=%s{extra}
        GROUP BY r.id, u.name, sem.name, ay.name
        ORDER BY r.created_at DESC
    """, tuple(params))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/guru/rooms/<room_id>/join', methods=['POST'])
@require_guru
def guru_join_room(room_id):
    """Guru bergabung ke room ujian."""
    room = query("SELECT id, name FROM rooms WHERE id=%s AND is_active=true AND school_id=%s",
                 (room_id, request.school_id), fetch='one')
    if not room: return jsonify({'error': 'Room tidak ditemukan'}), 404
    existing = query("SELECT 1 FROM room_teachers WHERE room_id=%s AND teacher_id=%s",
                     (room_id, request.user_id), fetch='one')
    if existing: return jsonify({'error': 'Sudah bergabung di room ini'}), 409
    query("INSERT INTO room_teachers (room_id, teacher_id) VALUES (%s,%s) ON CONFLICT DO NOTHING",
          (room_id, request.user_id), fetch='none')
    return jsonify({'ok': True, 'room_name': room['name']})

@admin_bp.route('/api/guru/rooms/<room_id>/leave', methods=['DELETE'])
@require_guru
def guru_leave_room(room_id):
    """Guru keluar dari room ujian."""
    query("DELETE FROM room_teachers WHERE room_id=%s AND teacher_id=%s",
          (room_id, request.user_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/guru/rooms/<room_id>/members', methods=['GET'])
@require_guru
def guru_room_members(room_id):
    """Guru bisa lihat siapa saja anggota room yang sudah dia join sendiri —
    TIDAK termasuk soal/isi ujian mereka, hanya nama."""
    is_member = query("SELECT 1 FROM room_teachers WHERE room_id=%s AND teacher_id=%s",
                      (room_id, request.user_id), fetch='one')
    if not is_member:
        return jsonify({'error': 'Anda belum join room ini'}), 403
    rows = query("""
        SELECT u.id, u.name, u.email FROM users u
        JOIN room_teachers rt ON rt.teacher_id=u.id
        WHERE rt.room_id=%s ORDER BY u.name
    """, (room_id,))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/guru/rooms/<room_id>/classes', methods=['GET'])
@require_guru
def guru_room_classes(room_id):
    """Guru bisa lihat kelas mana saja yang termasuk room ini (tidak pasti semua jenjang ikut)."""
    is_member = query("SELECT 1 FROM room_teachers WHERE room_id=%s AND teacher_id=%s",
                      (room_id, request.user_id), fetch='one')
    if not is_member:
        return jsonify({'error': 'Anda belum join room ini'}), 403
    rows = query("""
        SELECT c.* FROM classes c
        JOIN room_classes rc ON rc.class_id=c.id
        WHERE rc.room_id=%s ORDER BY c.grade, LENGTH(c.id), c.id
    """, (room_id,))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/guru/rooms/<room_id>/exams', methods=['GET'])
@require_guru
def guru_room_exams(room_id):
    """Guru bisa lihat siapa membuat ujian untuk jenjang apa di room ini —
    metadata saja (judul, mapel, jenjang, guru), TIDAK termasuk soal/kunci jawaban."""
    is_member = query("SELECT 1 FROM room_teachers WHERE room_id=%s AND teacher_id=%s",
                      (room_id, request.user_id), fetch='one')
    if not is_member:
        return jsonify({'error': 'Anda belum join room ini'}), 403
    rows = query("""
        SELECT e.id, e.title, e.grade, e.status, e.teacher_id,
               s.name as subject_name, u.name as teacher_name,
               sem.name as semester_name, ay.name as academic_year_name,
               (SELECT COUNT(*) FROM questions q WHERE q.exam_id=e.id) as question_count,
               (SELECT STRING_AGG(c.name,', ') FROM exam_classes ec
                JOIN classes c ON c.id=ec.class_id WHERE ec.exam_id=e.id) as class_names
        FROM exams e
        LEFT JOIN subjects s ON s.id=e.subject_id
        LEFT JOIN users u ON u.id=e.teacher_id
        LEFT JOIN semesters sem ON sem.id=e.semester_id
        LEFT JOIN academic_years ay ON ay.id=sem.academic_year_id
        WHERE e.room_id=%s
        ORDER BY e.grade, u.name
    """, (room_id,))
    return jsonify([dict(r) for r in rows])

# ══════════════════════════════════════════════════════════════
# GURU INVITATION SYSTEM — Hanya guru yang diundang admin bisa daftar
# ══════════════════════════════════════════════════════════════

@admin_bp.route('/api/admin/guru-invites', methods=['GET'])
@require_admin
def list_invites():
    rows = query("""
        SELECT gi.*, u.name as used_by_name
        FROM guru_invites gi
        LEFT JOIN users u ON u.id = gi.used_by
        WHERE gi.school_id=%s
        ORDER BY gi.created_at DESC
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

@admin_bp.route('/api/admin/guru-invites', methods=['POST'])
@require_admin
def create_invite():
    import secrets
    from datetime import datetime, timezone, timedelta
    body = request.json or {}
    email = body.get('email', '').strip().lower()
    name  = body.get('name', '').strip()
    if not email: return jsonify({'error': 'Email wajib diisi'}), 400

    # Cek apakah email sudah punya akun
    existing_user = query("SELECT id FROM users WHERE email=%s", (email,), fetch='one')
    if existing_user: return jsonify({'error': 'Email sudah terdaftar sebagai user'}), 409

    # Cek invite aktif
    existing = query("SELECT id FROM guru_invites WHERE email=%s AND used_at IS NULL AND expires_at > NOW()",
                     (email,), fetch='one')
    if existing: return jsonify({'error': 'Undangan aktif sudah ada untuk email ini'}), 409

    token = secrets.token_urlsafe(32)
    expires = datetime.now(timezone.utc) + timedelta(days=7)

    invite = query("""
        INSERT INTO guru_invites (id, email, name_hint, token, created_by, expires_at, school_id)
        VALUES (%s, %s, %s, %s, %s, %s, %s) RETURNING *
    """, (str(uuid.uuid4()), email, name, token, request.user_id, expires, request.school_id), fetch='one')

    # Return token untuk dikirim manual via email/WA
    # Pakai FRONTEND_URL dari env, fallback ke frontend_url dari request
    import os
    frontend_base = os.environ.get('FRONTEND_URL', body.get('frontend_url', '')).rstrip('/')
    invite_url = f"{frontend_base}/index.html?invite={token}&role=guru"
    return jsonify({**dict(invite), 'invite_url': invite_url, 'token': token}), 201

@admin_bp.route('/api/admin/guru-invites/<invite_id>', methods=['DELETE'])
@require_admin
def revoke_invite(invite_id):
    query("UPDATE guru_invites SET expires_at=NOW() WHERE id=%s AND school_id=%s",
          (invite_id, request.school_id), fetch='none')
    return jsonify({'ok': True})

@admin_bp.route('/api/auth/verify-invite', methods=['GET'])
def verify_invite():
    """Frontend cek apakah token invite valid sebelum login Google."""
    token = request.args.get('token', '')
    if not token: return jsonify({'valid': False, 'error': 'Token kosong'}), 400
    invite = query("""
        SELECT id, email, name_hint, expires_at, used_at
        FROM guru_invites WHERE token=%s
    """, (token,), fetch='one')
    if not invite: return jsonify({'valid': False, 'error': 'Token tidak ditemukan'}), 404
    if invite['used_at']: return jsonify({'valid': False, 'error': 'Token sudah dipakai'}), 400
    from datetime import datetime, timezone
    exp = invite['expires_at']
    if hasattr(exp, 'tzinfo') and exp.tzinfo is None: exp = exp.replace(tzinfo=timezone.utc)
    if exp < datetime.now(timezone.utc): return jsonify({'valid': False, 'error': 'Token kedaluwarsa'}), 400
    return jsonify({'valid': True, 'email': invite['email'], 'name_hint': invite['name_hint']})

# ── Device Management ──────────────────────────────────────────
@admin_bp.route('/api/admin/siswa/<siswa_id>/device', methods=['GET'])
@require_admin
def get_device_info(siswa_id):
    user = query("""
        SELECT id, name, email, device_id, device_info, last_login
        FROM users WHERE id=%s AND role='siswa' AND school_id=%s
    """, (siswa_id, request.school_id), fetch='one')
    if not user: return jsonify({'error': 'Siswa tidak ditemukan'}), 404
    return jsonify(dict(user))

@admin_bp.route('/api/admin/devices', methods=['GET'])
@require_admin
def list_devices():
    """Semua siswa + status device mereka."""
    rows = query("""
        SELECT u.id, u.name, u.email, u.class_id, c.name as class_name,
               u.device_id, u.device_info, u.last_login,
               CASE WHEN u.device_id IS NOT NULL THEN true ELSE false END as has_device
        FROM users u
        LEFT JOIN classes c ON c.id=u.class_id
        WHERE u.role='siswa' AND u.school_id=%s
        ORDER BY c.name, u.name
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])

# ── Guru invite check di Google callback ──────────────────────
# (dipatch di app.py google_callback untuk validasi)
def check_guru_invite(email, token):
    """Return invite record kalau valid, None kalau tidak."""
    if not token: return None
    invite = query("""
        SELECT * FROM guru_invites
        WHERE token=%s AND email=%s AND used_at IS NULL AND expires_at > NOW()
    """, (token, email.lower()), fetch='one')
    return invite

def mark_invite_used(token, user_id):
    query("UPDATE guru_invites SET used_at=NOW(), used_by=%s WHERE token=%s",
          (user_id, token), fetch='none')

# ── Settings CRUD (alias ke exam-settings dengan schema kolom) ──
@admin_bp.route('/api/admin/settings', methods=['GET'])
@require_admin
def get_settings():
    s = query("SELECT * FROM exam_settings WHERE school_id=%s LIMIT 1", (request.school_id,), fetch='one')
    return jsonify(dict(s) if s else {
        'passing_grade': 75, 'allow_remedial': True,
        'max_violations': 5, 'auto_submit_on_violation': True, 'show_ranking': True,
        'proctor_code': None
    })

@admin_bp.route('/api/admin/settings', methods=['POST'])
@require_admin
def save_settings():
    _ALLOWED = {'passing_grade', 'allow_remedial', 'max_violations',
                'auto_submit_on_violation', 'show_ranking', 'proctor_code'}
    data = {k: v for k, v in (request.json or {}).items() if k in _ALLOWED}
    if not data:
        return jsonify({'error': 'Tidak ada field yang valid'}), 400
    existing = query("SELECT id FROM exam_settings WHERE school_id=%s LIMIT 1",
                     (request.school_id,), fetch='one')
    if existing:
        fields = ', '.join(f"{k}=%s" for k in data)
        query(f"UPDATE exam_settings SET {fields} WHERE id=%s",
              list(data.values()) + [str(existing['id'])], fetch='none')
    else:
        cols = ', '.join(list(data.keys()) + ['school_id'])
        vals = ', '.join(['%s'] * (len(data) + 1))
        query(f"INSERT INTO exam_settings ({cols}) VALUES ({vals})",
              list(data.values()) + [request.school_id], fetch='none')
    return jsonify({'ok': True})

# ── Reset ALL devices ────────────────────────────────────────
@admin_bp.route('/api/admin/devices/reset-all', methods=['POST'])
@require_admin
def reset_all_devices():
    query("UPDATE users SET device_id=NULL, device_info=NULL WHERE role='siswa' AND school_id=%s",
          (request.school_id,), fetch='none')
    return jsonify({'ok': True})

# ══════════════════════════════════════════════════════════════
# SESSION REOPEN — Opsi 2: Guru/Admin buka ulang sesi siswa
# ══════════════════════════════════════════════════════════════

def _do_reopen(session_id, school_id, extra_min=None, keep_timer=False, reset_answers=False, reset_violations=False):
    """Helper internal: buka ulang sesi. Dipakai oleh semua endpoint reopen."""
    from datetime import datetime, timezone, timedelta

    sess = query("""SELECT es.*, e.teacher_id, e.duration_minutes
                    FROM exam_sessions es JOIN exams e ON e.id = es.exam_id
                    WHERE es.id = %s AND e.school_id = %s""", (session_id, school_id), fetch='one')
    if not sess:
        return None, 'Sesi tidak ditemukan'

    if reset_answers:
        query("DELETE FROM answers       WHERE session_id=%s", (session_id,), fetch='none')
        query("DELETE FROM results       WHERE session_id=%s", (session_id,), fetch='none')
        try:
            query("DELETE FROM essay_answers WHERE session_id=%s", (session_id,), fetch='none')
            query("DELETE FROM multi_answers  WHERE session_id=%s", (session_id,), fetch='none')
        except Exception:
            pass  # tabel mungkin belum ada

    now = datetime.now(timezone.utc)

    if keep_timer:
        # Lanjutkan countdown asli — jangan ubah expires_at
        new_expires = sess.get('expires_at')
        # Jika expires_at sudah lewat, perpanjang dengan durasi ujian asli
        if not new_expires or (hasattr(new_expires, 'tzinfo') and new_expires < now):
            dur = int(sess.get('duration_minutes') or 90)
            new_expires = now + timedelta(minutes=dur)
        update_expires_sql = "expires_at = %s,"
        update_expires_val = (new_expires,)
    else:
        dur = int(extra_min or 30)
        new_expires = now + timedelta(minutes=dur)
        update_expires_sql = "expires_at = %s,"
        update_expires_val = (new_expires,)

    viol_reset = ", tab_violations = 0" if reset_violations else ""

    query(f"""UPDATE exam_sessions
              SET submitted_at  = NULL,
                  auto_submitted = FALSE,
                  status         = 'ongoing',
                  exit_allowed   = FALSE,
                  {update_expires_sql}
                  started_at     = COALESCE(started_at, %s)
                  {viol_reset}
              WHERE id = %s""",
          (*update_expires_val, now, session_id), fetch='none')

    return sess, None

@admin_bp.route('/api/admin/sessions/<session_id>/reopen', methods=['POST'])
@require_admin
def reopen_session(session_id):
    body      = request.json or {}
    extra_min = int(body.get('extra_minutes', 15))
    keep      = bool(body.get('keep_timer', False))
    sess, err = _do_reopen(session_id, request.school_id, extra_min=extra_min, keep_timer=keep)
    if err: return jsonify({'error': err}), 404
    log_activity(request.user_id, 'SESSION_REOPEN',
                 f"Buka ulang sesi {session_id[:8]}", request.remote_addr, request.school_id)
    return jsonify({'ok': True})

@admin_bp.route('/api/guru/sessions/<session_id>/reopen', methods=['POST'])
@require_guru
def guru_reopen_session(session_id):
    """Lanjutkan atau perpanjang sesi — tanpa reset jawaban."""
    body      = request.json or {}
    extra_min = int(body.get('extra_minutes', 30))
    keep      = bool(body.get('keep_timer', False))
    sess, err = _do_reopen(session_id, request.school_id, extra_min=extra_min, keep_timer=keep)
    if err: return jsonify({'error': err}), 404
    if request.user_role != 'admin' and str(sess['teacher_id']) != request.user_id:
        return jsonify({'error': 'Akses ditolak'}), 403
    return jsonify({'ok': True})

@admin_bp.route('/api/guru/sessions/<session_id>/reset-reopen', methods=['POST'])
@require_guru
def guru_reset_reopen(session_id):
    """Reset semua jawaban + buka ulang sesi (mulai dari awal)."""
    body      = request.json or {}
    extra_min = int(body.get('extra_minutes', 90))
    sess, err = _do_reopen(session_id, request.school_id, extra_min=extra_min,
                           reset_answers=True, reset_violations=True)
    if err: return jsonify({'error': err}), 404
    if request.user_role != 'admin' and str(sess['teacher_id']) != request.user_id:
        return jsonify({'error': 'Akses ditolak'}), 403
    log_activity(request.user_id, 'SESSION_RESET',
                 f"Reset jawaban sesi {session_id[:8]}", request.remote_addr, request.school_id)
    return jsonify({'ok': True})

# ── Cleanup finished exams ───────────────────────────────────
@admin_bp.route('/api/admin/exams/cleanup', methods=['DELETE'])
@require_admin
def cleanup_exams():
    query("DELETE FROM exams WHERE status='finished' AND school_id=%s", (request.school_id,), fetch='none')
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════
# ACTIVITY LOG — log otomatis saat aksi penting terjadi
# Tabel: CREATE TABLE IF NOT EXISTS activity_logs (
#   id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
#   user_id UUID REFERENCES users(id) ON DELETE SET NULL,
#   action VARCHAR(64), detail TEXT, ip_address VARCHAR(64),
#   created_at TIMESTAMPTZ DEFAULT NOW()
# );
# ══════════════════════════════════════════════════════════════

def _ensure_log_table():
    try:
        query("""CREATE TABLE IF NOT EXISTS activity_logs (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            user_id UUID REFERENCES users(id) ON DELETE SET NULL,
            action VARCHAR(64) NOT NULL,
            detail TEXT,
            ip_address VARCHAR(64),
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""", fetch='none')
    except Exception:
        pass

_ensure_log_table()


@admin_bp.route('/api/admin/activity-logs', methods=['GET'])
@require_admin
def get_activity_logs():
    limit  = min(int(request.args.get('limit', 50)), 200)
    action = request.args.get('action', '')
    where  = "WHERE al.school_id=%s"
    params = [request.school_id]
    if action:
        where += " AND al.action=%s"; params.append(action)
    else:
        # Reset/buka-ulang sesi siswa terlalu sering terjadi saat guru aktif
        # mengoreksi ujian — kalau ikut ditampilkan, banjir log ini menutupi
        # aktivitas lain yang lebih relevan (login, buat ujian/soal, dst).
        # Tetap bisa diambil lewat ?action=SESSION_RESET kalau memang perlu.
        where += " AND al.action NOT IN ('SESSION_RESET','SESSION_REOPEN')"
    try:
        rows = query(f"""
            SELECT al.id, al.action, al.detail, al.ip_address, al.created_at,
                   u.name as user_name, u.role as user_role
            FROM activity_logs al
            LEFT JOIN users u ON u.id = al.user_id
            {where}
            ORDER BY al.created_at DESC
            LIMIT %s
        """, params + [limit])
        return jsonify([dict(r) for r in rows])
    except Exception:
        return jsonify([])


@admin_bp.route('/api/admin/login-history', methods=['GET'])
@require_admin
def login_history():
    limit = min(int(request.args.get('limit', 30)), 100)
    try:
        rows = query("""
            SELECT u.id, u.name, u.email, u.role, u.last_login, u.avatar_url
            FROM users u
            WHERE u.last_login IS NOT NULL AND u.school_id=%s
            ORDER BY u.last_login DESC
            LIMIT %s
        """, (request.school_id, limit))
        return jsonify([dict(r) for r in rows])
    except Exception:
        return jsonify([])


# ══════════════════════════════════════════════════════════════
# ANALYTICS — Statistik sekolah untuk dashboard
# ══════════════════════════════════════════════════════════════

@admin_bp.route('/api/admin/analytics', methods=['GET'])
@require_admin
def get_analytics():
    from datetime import datetime, timezone, timedelta
    sid = request.school_id

    total_guru  = query("SELECT COUNT(*) as n FROM users WHERE role='guru' AND school_id=%s", (sid,), fetch='one')['n']
    total_siswa = query("SELECT COUNT(*) as n FROM users WHERE role='siswa' AND school_id=%s", (sid,), fetch='one')['n']
    total_kelas = query("SELECT COUNT(*) as n FROM classes WHERE school_id=%s", (sid,), fetch='one')['n']
    total_ujian = query("SELECT COUNT(*) as n FROM exams WHERE school_id=%s", (sid,), fetch='one')['n']
    ujian_aktif = query("SELECT COUNT(*) as n FROM exams WHERE status IN ('published','ongoing') AND school_id=%s", (sid,), fetch='one')['n']
    device_ok   = query("SELECT COUNT(*) as n FROM users WHERE role='siswa' AND device_id IS NOT NULL AND school_id=%s", (sid,), fetch='one')['n']

    # Tingkat kelulusan
    settings_row = query("SELECT passing_grade FROM exam_settings WHERE school_id=%s LIMIT 1", (sid,), fetch='one')
    passing_grade = float(settings_row['passing_grade']) if settings_row and settings_row.get('passing_grade') else 75.0
    total_results = query("SELECT COUNT(*) as n FROM results r JOIN exam_sessions es ON es.id=r.session_id WHERE es.submitted_at IS NOT NULL AND es.school_id=%s", (sid,), fetch='one')['n']
    lulus_count   = query("SELECT COUNT(*) as n FROM results r JOIN exam_sessions es ON es.id=r.session_id WHERE es.submitted_at IS NOT NULL AND es.school_id=%s AND r.score >= %s", (sid, passing_grade), fetch='one')['n']
    pass_rate = round((lulus_count / total_results * 100), 1) if total_results else 0

    # Rata-rata nilai keseluruhan
    avg_row = query("SELECT ROUND(AVG(r.score)::numeric, 1) as avg FROM results r JOIN exam_sessions es ON es.id=r.session_id WHERE es.submitted_at IS NOT NULL AND es.school_id=%s", (sid,), fetch='one')
    avg_score = float(avg_row['avg']) if avg_row and avg_row['avg'] else 0

    # Distribusi nilai (A≥90, B≥75, C≥60, D<60)
    dist = query("""
        SELECT
            COUNT(*) FILTER (WHERE r.score >= 90) as a,
            COUNT(*) FILTER (WHERE r.score >= 75 AND r.score < 90) as b,
            COUNT(*) FILTER (WHERE r.score >= 60 AND r.score < 75) as c,
            COUNT(*) FILTER (WHERE r.score < 60) as d
        FROM results r JOIN exam_sessions es ON es.id=r.session_id
        WHERE es.submitted_at IS NOT NULL AND es.school_id=%s
    """, (sid,), fetch='one')

    # Ujian bulan ini
    ujian_bulan_ini = query("""
        SELECT COUNT(*) as n FROM exams
        WHERE created_at >= date_trunc('month', NOW()) AND school_id=%s
    """, (sid,), fetch='one')['n']

    # Guru pending
    guru_pending = query("SELECT COUNT(*) as n FROM users WHERE role='guru_pending' AND school_id=%s", (sid,), fetch='one')['n']

    return jsonify({
        'total_guru': total_guru,
        'total_siswa': total_siswa,
        'total_kelas': total_kelas,
        'total_ujian': total_ujian,
        'ujian_aktif': ujian_aktif,
        'device_terdaftar': device_ok,
        'pass_rate': pass_rate,
        'avg_score': avg_score,
        'grade_distribution': {
            'A': int(dist['a'] or 0), 'B': int(dist['b'] or 0),
            'C': int(dist['c'] or 0), 'D': int(dist['d'] or 0),
        },
        'ujian_bulan_ini': ujian_bulan_ini,
        'guru_pending': guru_pending,
        'passing_grade': passing_grade,
    })


@admin_bp.route('/api/admin/analytics/grades-by-subject', methods=['GET'])
@require_admin
def grades_by_subject():
    rows = query("""
        SELECT s.name as subject_name,
               ROUND(AVG(r.score)::numeric, 1) as avg_score,
               COUNT(r.id) as total_sessions,
               COUNT(*) FILTER (WHERE r.score >= 75) as lulus
        FROM results r
        JOIN exam_sessions es ON es.id = r.session_id
        JOIN exams e ON e.id = es.exam_id
        JOIN subjects s ON s.id = e.subject_id
        WHERE es.submitted_at IS NOT NULL AND e.school_id=%s
        GROUP BY s.id, s.name
        ORDER BY avg_score DESC
        LIMIT 15
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])


@admin_bp.route('/api/admin/analytics/exam-trend', methods=['GET'])
@require_admin
def exam_trend():
    rows = query("""
        SELECT TO_CHAR(DATE_TRUNC('month', created_at), 'YYYY-MM') as month,
               COUNT(*) as total
        FROM exams
        WHERE created_at >= NOW() - INTERVAL '6 months' AND school_id=%s
        GROUP BY DATE_TRUNC('month', created_at)
        ORDER BY DATE_TRUNC('month', created_at)
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])


@admin_bp.route('/api/admin/exam-monitor', methods=['GET'])
@require_admin
def exam_monitor():
    """Sesi ujian aktif saat ini beserta info device siswa."""
    rows = query("""
        SELECT es.id as session_id, es.exam_id, es.student_id, es.status,
               es.started_at, es.expires_at, es.tab_violations,
               es.ip_address, es.device_key,
               u.name as student_name, u.class_id, u.device_id,
               c.name as class_name,
               e.title as exam_title,
               (SELECT COUNT(*) FROM answers a WHERE a.session_id=es.id) as answered_count,
               (SELECT COUNT(*) FROM questions q WHERE q.exam_id=es.exam_id) as total_questions
        FROM exam_sessions es
        JOIN users u ON u.id = es.student_id
        LEFT JOIN classes c ON c.id = u.class_id
        JOIN exams e ON e.id = es.exam_id
        WHERE es.submitted_at IS NULL AND es.status='ongoing' AND es.school_id=%s
        ORDER BY es.started_at DESC
    """, (request.school_id,))
    return jsonify([dict(r) for r in rows])


# ══════════════════════════════════════════════════════════════
# TAHUN AJARAN & SEMESTER
# Tabel:
#   CREATE TABLE IF NOT EXISTS academic_years (
#     id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
#     name VARCHAR(20) NOT NULL UNIQUE,
#     is_active BOOLEAN DEFAULT false,
#     start_date DATE, end_date DATE,
#     created_at TIMESTAMPTZ DEFAULT NOW()
#   );
#   CREATE TABLE IF NOT EXISTS semesters (
#     id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
#     academic_year_id UUID REFERENCES academic_years(id) ON DELETE CASCADE,
#     name VARCHAR(20) NOT NULL,
#     is_active BOOLEAN DEFAULT false,
#     start_date DATE, end_date DATE,
#     created_at TIMESTAMPTZ DEFAULT NOW()
#   );
# ══════════════════════════════════════════════════════════════

def _ensure_year_tables():
    try:
        query("""CREATE TABLE IF NOT EXISTS academic_years (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            name VARCHAR(20) NOT NULL UNIQUE,
            is_active BOOLEAN DEFAULT false,
            start_date DATE, end_date DATE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""", fetch='none')
        query("""CREATE TABLE IF NOT EXISTS semesters (
            id UUID PRIMARY KEY DEFAULT gen_random_uuid(),
            academic_year_id UUID REFERENCES academic_years(id) ON DELETE CASCADE,
            name VARCHAR(20) NOT NULL,
            is_active BOOLEAN DEFAULT false,
            start_date DATE, end_date DATE,
            created_at TIMESTAMPTZ DEFAULT NOW()
        )""", fetch='none')
    except Exception:
        pass

_ensure_year_tables()


@admin_bp.route('/api/admin/academic-years', methods=['GET'])
@require_admin
def list_academic_years():
    try:
        rows = query("""
            SELECT ay.*,
                   COUNT(s.id) as semester_count,
                   COUNT(s.id) FILTER (WHERE s.is_active) as active_semesters,
                   (SELECT COUNT(*) FROM rooms r WHERE r.semester_id IN
                       (SELECT id FROM semesters WHERE academic_year_id=ay.id)) as room_count,
                   (SELECT COUNT(*) FROM exams e WHERE e.semester_id IN
                       (SELECT id FROM semesters WHERE academic_year_id=ay.id)) as exam_count,
                   (SELECT ROUND(AVG(res.score)::numeric, 1) FROM results res
                    JOIN exams e2 ON e2.id=res.exam_id
                    WHERE e2.semester_id IN (SELECT id FROM semesters WHERE academic_year_id=ay.id)) as avg_score
            FROM academic_years ay
            LEFT JOIN semesters s ON s.academic_year_id = ay.id
            WHERE ay.school_id=%s
            GROUP BY ay.id ORDER BY ay.name DESC
        """, (request.school_id,))
        return jsonify([dict(r) for r in rows])
    except Exception:
        # Fallback kalau kolom semester_id di rooms/exams belum ada (migrasi
        # belum dijalankan) — tetap tampilkan data inti, jangan sampai
        # tahun ajaran yang sudah ada terlihat hilang.
        try:
            rows = query("""
                SELECT ay.*,
                       COUNT(s.id) as semester_count,
                       COUNT(s.id) FILTER (WHERE s.is_active) as active_semesters
                FROM academic_years ay
                LEFT JOIN semesters s ON s.academic_year_id = ay.id
                WHERE ay.school_id=%s
                GROUP BY ay.id ORDER BY ay.name DESC
            """, (request.school_id,))
            return jsonify([dict(r) for r in rows])
        except Exception:
            return jsonify([])


@admin_bp.route('/api/admin/academic-years', methods=['POST'])
@require_admin
def create_academic_year():
    body = request.json or {}
    name = body.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Nama tahun ajaran wajib'}), 400
    try:
        row = query("""
            INSERT INTO academic_years (id, name, start_date, end_date, is_active, school_id)
            VALUES (gen_random_uuid(), %s, %s, %s, %s, %s) RETURNING *
        """, (name, body.get('start_date'), body.get('end_date'), False, request.school_id), fetch='one')
        log_activity(request.user_id, 'TAHUN_AJARAN_BUAT', f"Buat tahun ajaran {name}", request.remote_addr, request.school_id)
        return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@admin_bp.route('/api/admin/academic-years/<year_id>', methods=['PATCH'])
@require_admin
def update_academic_year(year_id):
    body = request.json or {}
    if body.get('is_active'):
        query("UPDATE academic_years SET is_active=false WHERE school_id=%s", (request.school_id,), fetch='none')
    fields, vals = [], []
    for f in ['name', 'start_date', 'end_date', 'is_active']:
        if f in body:
            fields.append(f"{f}=%s"); vals.append(body[f])
    if fields:
        query(f"UPDATE academic_years SET {', '.join(fields)} WHERE id=%s AND school_id=%s",
              vals + [year_id, request.school_id], fetch='none')
    return jsonify({'ok': True})


@admin_bp.route('/api/admin/academic-years/<year_id>', methods=['DELETE'])
@require_admin
def delete_academic_year(year_id):
    query("DELETE FROM academic_years WHERE id=%s AND school_id=%s", (year_id, request.school_id), fetch='none')
    return jsonify({'ok': True})


@admin_bp.route('/api/admin/semesters', methods=['GET'])
@require_admin
def list_semesters():
    year_id = request.args.get('year_id', '')
    stats_select = """,
                   (SELECT COUNT(*) FROM rooms r WHERE r.semester_id=s.id) as room_count,
                   (SELECT COUNT(*) FROM exams e WHERE e.semester_id=s.id) as exam_count,
                   (SELECT ROUND(AVG(res.score)::numeric, 1) FROM results res
                    JOIN exams e2 ON e2.id=res.exam_id WHERE e2.semester_id=s.id) as avg_score"""
    try:
        if year_id:
            rows = query(f"""
                SELECT s.*, ay.name as year_name{stats_select} FROM semesters s
                JOIN academic_years ay ON ay.id = s.academic_year_id
                WHERE s.academic_year_id=%s AND ay.school_id=%s ORDER BY s.start_date
            """, (year_id, request.school_id))
        else:
            rows = query(f"""
                SELECT s.*, ay.name as year_name{stats_select} FROM semesters s
                JOIN academic_years ay ON ay.id = s.academic_year_id
                WHERE ay.school_id=%s
                ORDER BY ay.name DESC, s.start_date
            """, (request.school_id,))
        return jsonify([dict(r) for r in rows])
    except Exception:
        # Fallback kalau kolom semester_id di rooms/exams belum ada
        try:
            if year_id:
                rows = query("""
                    SELECT s.*, ay.name as year_name FROM semesters s
                    JOIN academic_years ay ON ay.id = s.academic_year_id
                    WHERE s.academic_year_id=%s AND ay.school_id=%s ORDER BY s.start_date
                """, (year_id, request.school_id))
            else:
                rows = query("""
                    SELECT s.*, ay.name as year_name FROM semesters s
                    JOIN academic_years ay ON ay.id = s.academic_year_id
                    WHERE ay.school_id=%s
                    ORDER BY ay.name DESC, s.start_date
                """, (request.school_id,))
            return jsonify([dict(r) for r in rows])
        except Exception:
            return jsonify([])


@admin_bp.route('/api/admin/semesters', methods=['POST'])
@require_admin
def create_semester():
    body = request.json or {}
    year_id = body.get('academic_year_id', '').strip()
    name    = body.get('name', '').strip()
    if not year_id or not name:
        return jsonify({'error': 'academic_year_id dan name wajib'}), 400
    year = query("SELECT id FROM academic_years WHERE id=%s AND school_id=%s", (year_id, request.school_id), fetch='one')
    if not year:
        return jsonify({'error': 'Tahun ajaran tidak ditemukan'}), 404
    try:
        row = query("""
            INSERT INTO semesters (id, academic_year_id, name, start_date, end_date, is_active, school_id)
            VALUES (gen_random_uuid(), %s, %s, %s, %s, %s, %s) RETURNING *
        """, (year_id, name, body.get('start_date'), body.get('end_date'), False, request.school_id), fetch='one')
        return jsonify(dict(row)), 201
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@admin_bp.route('/api/admin/semesters/<sem_id>', methods=['PATCH'])
@require_admin
def update_semester(sem_id):
    body = request.json or {}
    if body.get('is_active'):
        sem = query("SELECT academic_year_id FROM semesters WHERE id=%s AND school_id=%s",
                    (sem_id, request.school_id), fetch='one')
        if sem:
            query("UPDATE semesters SET is_active=false WHERE academic_year_id=%s",
                  (sem['academic_year_id'],), fetch='none')
    fields, vals = [], []
    for f in ['name', 'start_date', 'end_date', 'is_active']:
        if f in body:
            fields.append(f"{f}=%s"); vals.append(body[f])
    if fields:
        query(f"UPDATE semesters SET {', '.join(fields)} WHERE id=%s AND school_id=%s",
              vals + [sem_id, request.school_id], fetch='none')
    return jsonify({'ok': True})


@admin_bp.route('/api/admin/semesters/<sem_id>', methods=['DELETE'])
@require_admin
def delete_semester(sem_id):
    query("DELETE FROM semesters WHERE id=%s AND school_id=%s", (sem_id, request.school_id), fetch='none')
    return jsonify({'ok': True})


@admin_bp.route('/api/admin/active-period', methods=['GET'])
@require_admin
def get_active_period():
    """Tahun ajaran dan semester yang sedang aktif."""
    try:
        year = query("SELECT * FROM academic_years WHERE is_active=true AND school_id=%s LIMIT 1",
                     (request.school_id,), fetch='one')
        sem  = query("SELECT * FROM semesters WHERE is_active=true AND school_id=%s LIMIT 1",
                     (request.school_id,), fetch='one')
        return jsonify({
            'academic_year': dict(year) if year else None,
            'semester': dict(sem) if sem else None,
        })
    except Exception:
        return jsonify({'academic_year': None, 'semester': None})